import hashlib
import importlib
import json
import os
import re
import sqlite3
import sys
import unicodedata
import uuid
from email import policy
from email.parser import BytesParser
from datetime import datetime, timezone
from pathlib import Path

from fastapi import HTTPException, UploadFile

from app.core.correlation import get_correlation_id
from app.core.queue import enqueue_document_processing
from app.months import month_number, parse_issue_date
from app.repositories.document_repository import get_db_connection
from app.schemas.documents import (
    OverrideRequest,
    RenditionGenerateByFilterRequest,
    RenditionGenerateRequest,
    ReviewActionRequest,
    ReviewOverridesRequest,
    ReviewResolveRequest,
)

EXPORT_DIR = os.getenv("EXPORT_DIR", "/tmp/capturador_exports")
STORAGE_DIR = os.getenv("STORAGE_DIR", "/tmp/capturador_uploads")
TEMPLATE_VERSION_DEFAULT = "01-rendicion-gastos-2025"
COMPANY_RUT_DEFAULT = os.getenv("COMPANY_RUT", "76000000-0")
DEFAULT_APPROVER = os.getenv("DEFAULT_APPROVER", "PENDIENTE")
DEFAULT_CENTER_COST = os.getenv("DEFAULT_CENTER_COST", "GENERAL")
DEFAULT_ACCOUNT = os.getenv("DEFAULT_ACCOUNT", "510100")
DEFAULT_ACCOUNT_DESCRIPTION = os.getenv("DEFAULT_ACCOUNT_DESCRIPTION", "Gastos Operacionales")
TEMPLATE_XLSX_PATH = os.getenv(
    "RENDITION_TEMPLATE_PATH",
    str(Path(__file__).resolve().parents[3] / "docs" / "01 Rendicion de gastos 2025.xlsx"),
)
MAX_UPLOAD_BYTES = 10 * 1024 * 1024
ALLOWED_MIME_TYPES = {
    "application/pdf",
    "image/jpeg",
    "image/png",
    "message/rfc822",
}
ALLOWED_EXTENSIONS = {".pdf", ".jpg", ".jpeg", ".png", ".eml"}
EXPECTED_MIME_BY_EXTENSION = {
    ".pdf": "application/pdf",
    ".jpg": "image/jpeg",
    ".jpeg": "image/jpeg",
    ".png": "image/png",
    ".eml": "message/rfc822",
}
UPLOAD_QUARANTINE_ENABLED = os.getenv("UPLOAD_QUARANTINE_ENABLED", "false").lower() in {"1", "true", "yes", "on"}
UPLOAD_QUARANTINE_DIR = os.getenv("UPLOAD_QUARANTINE_DIR", str(Path(STORAGE_DIR).parent / "capturador_quarantine"))

HEADER_FIELDS = [
    "Responsable",
    "Rut",
    "Periodo",
    "Monto a pagar",
    "Fecha",
    "Autorizado por",
]

DETAIL_FIELDS = [
    "Mes",
    "Cuenta",
    "Descripcion Cuenta",
    "Fecha",
    "Nro Boleta Factura",
    "Descripcion del gasto",
    "Concepto",
    "Centro Costo",
    "PROVEEDOR",
    "RUT",
    "Observaciones",
    "Total",
]


def _utc_now_iso() -> str:
    return datetime.now(tz=timezone.utc).isoformat()


def _utc_today() -> str:
    return datetime.now(tz=timezone.utc).date().isoformat()


def _add_warning(warnings: list[dict[str, str]], field_name: str) -> None:
    warnings.append(
        {
            "code": "MISSING_FIELD",
            "field": field_name,
            "severity": "warning",
            "message": "No se detecto valor para el campo; se exporta en blanco",
        }
    )


def _add_custom_warning(
    warnings: list[dict[str, str]],
    *,
    code: str,
    field_name: str,
    message: str,
    severity: str = "warning",
) -> None:
    warnings.append(
        {
            "code": code,
            "field": field_name,
            "severity": severity,
            "message": message,
        }
    )


def _normalize_currency(raw_text: str) -> str:
    if re.search(r"\b(clp|peso|pesos|\$)\b", raw_text, flags=re.IGNORECASE):
        return "CLP"
    if re.search(r"\b(usd|us\$|dolar|dolares)\b", raw_text, flags=re.IGNORECASE):
        return "USD"
    if re.search(r"\b(eur|euro)\b", raw_text, flags=re.IGNORECASE):
        return "EUR"
    return "CLP"


def _parse_amount(raw_value: str) -> float | None:
    value = raw_value.strip()
    value = re.sub(r"[^0-9,.-]", "", value)
    if not value:
        return None
    if value.count(",") > 0 and value.count(".") > 0:
        value = value.replace(".", "").replace(",", ".")
    elif value.count(",") > 0:
        value = value.replace(".", "").replace(",", ".")
    else:
        value = value.replace(",", "")
    try:
        return float(value)
    except ValueError:
        return None


def _format_rut(raw_rut: str) -> str:
    rut = raw_rut.replace(".", "").strip().upper()
    if "-" in rut:
        return rut
    if len(rut) >= 2:
        return f"{rut[:-1]}-{rut[-1]}"
    return rut


def _sanitize_filename(raw_name: str | None) -> str:
    name = Path(raw_name or "uploaded_file").name
    clean = re.sub(r"[^A-Za-z0-9._\- ]", "_", name).strip()
    clean = re.sub(r"\s+", " ", clean)
    return clean or "uploaded_file"


def _detect_file_signature(content: bytes) -> str | None:
    if content.startswith(b"%PDF"):
        return ".pdf"
    if content.startswith(b"\x89PNG\r\n\x1a\n"):
        return ".png"
    if content.startswith(b"\xff\xd8\xff"):
        return ".jpg"
    # For EML we allow plain text-like payloads and rely on parser later.
    if b"\n" in content[:2048] and (b"Subject:" in content[:2048] or b"From:" in content[:2048]):
        return ".eml"
    return None


def _canonical_extension(extension: str | None) -> str:
    ext = (extension or "").lower().strip()
    if ext == ".jpeg":
        return ".jpg"
    return ext


def _audit_upload_blocked(*, tenant_id: str, file_name: str, reason: str) -> None:
    conn = get_db_connection()
    try:
        _persist_audit_event(
            conn,
            entity_type="upload",
            entity_id=str(uuid.uuid4()),
            event_type="DOCUMENT_UPLOAD_BLOCKED",
            payload={"tenant_id": tenant_id, "file_name": file_name, "reason": reason},
        )
        conn.commit()
    finally:
        conn.close()


def _quarantine_upload(*, safe_file_name: str, content: bytes, reason: str) -> str:
    quarantine_dir = Path(UPLOAD_QUARANTINE_DIR)
    quarantine_dir.mkdir(parents=True, exist_ok=True)
    quarantine_name = f"{uuid.uuid4()}_{safe_file_name}"
    quarantine_path = quarantine_dir / quarantine_name
    quarantine_path.write_bytes(content)

    sidecar = quarantine_path.with_suffix(quarantine_path.suffix + ".json")
    sidecar.write_text(json.dumps({"reason": reason, "created_at": _utc_now_iso()}, ensure_ascii=True), encoding="utf-8")
    return str(quarantine_path)


def _extract_text_from_pdf(storage_path: str | None) -> str:
    if not storage_path:
        return ""
    file_path = Path(storage_path)
    if not file_path.exists():
        return ""
    try:
        from pypdf import PdfReader
    except ModuleNotFoundError:
        return ""

    try:
        reader = PdfReader(str(file_path))
        parts: list[str] = []
        for page in reader.pages:
            parts.append(page.extract_text() or "")
        return "\n".join(parts)
    except Exception:
        return ""


def _extract_text_from_eml(storage_path: str | None) -> str:
    if not storage_path:
        return ""
    file_path = Path(storage_path)
    if not file_path.exists():
        return ""

    try:
        raw_bytes = file_path.read_bytes()
        message = BytesParser(policy=policy.default).parsebytes(raw_bytes)
    except Exception:
        return ""

    parts: list[str] = []
    subject = message.get("subject")
    sender = message.get("from")
    if subject:
        parts.append(f"Subject: {subject}")
    if sender:
        parts.append(f"From: {sender}")

    try:
        if message.is_multipart():
            for part in message.walk():
                content_type = part.get_content_type()
                if content_type == "text/plain":
                    payload = part.get_content()
                    if payload:
                        parts.append(str(payload))
        else:
            payload = message.get_content()
            if payload:
                parts.append(str(payload))
    except Exception:
        pass

    return "\n".join(parts)


def _load_back_ocr_module():
    configured = os.getenv("BACK_CODE_PATH")
    if configured:
        back_dir = Path(configured)
    else:
        back_dir = Path(__file__).resolve().parents[3] / "back"
        if not back_dir.exists():
            back_dir = Path("/opt/back")

    back_path = str(back_dir)
    if back_path not in sys.path:
        sys.path.insert(0, back_path)

    try:
        return importlib.import_module("app.ocr_pipeline")
    except ModuleNotFoundError:
        return None


def _load_back_intelligence_module():
    configured = os.getenv("BACK_CODE_PATH")
    if configured:
        back_dir = Path(configured)
    else:
        back_dir = Path(__file__).resolve().parents[3] / "back"
        if not back_dir.exists():
            back_dir = Path("/opt/back")

    back_path = str(back_dir)
    if back_path not in sys.path:
        sys.path.insert(0, back_path)

    try:
        return importlib.import_module("app.intelligent_extraction")
    except ModuleNotFoundError:
        return None


def _load_back_ollama_module():
    configured = os.getenv("BACK_CODE_PATH")
    if configured:
        back_dir = Path(configured)
    else:
        back_dir = Path(__file__).resolve().parents[3] / "back"
        if not back_dir.exists():
            back_dir = Path("/opt/back")

    back_path = str(back_dir)
    if back_path not in sys.path:
        sys.path.insert(0, back_path)

    try:
        return importlib.import_module("app.ollama_adapter")
    except ModuleNotFoundError:
        return None


def _infer_document_type(raw_text: str, file_name: str) -> tuple[str, float]:
    text = f"{raw_text}\n{file_name}".lower()
    if "google play" in text or "youtube" in text or "recibo" in text or "receipt" in text:
        return "receipt", 0.87
    if "invoice" in text:
        return "invoice", 0.91
    if "factura" in text:
        return "invoice", 0.93
    if "boleta" in text or "ticket" in text:
        return "receipt", 0.92
    if "credit note" in text or "nota de credito" in text:
        return "credit_note", 0.88
    return "unknown", 0.55


def _clean_text(raw_text: str) -> str:
    return "".join(ch for ch in raw_text if ch.isprintable() or ch in {"\n", "\r", "\t"})


def _extract_supplier_name(raw_text: str) -> str | None:
    lines = [line.strip() for line in raw_text.splitlines() if line.strip()]
    if not lines:
        return None

    for idx, line in enumerate(lines):
        if re.search(r"\bcl\s*tin\b", line, flags=re.IGNORECASE):
            for back in range(idx - 1, max(-1, idx - 5), -1):
                candidate = lines[back]
                if re.search(r"\d", candidate):
                    continue
                if candidate.lower() in {"bill to", "description", "invoice", "openai"}:
                    continue
                if len(candidate) < 3:
                    continue
                return candidate

    m = re.search(r"(?:razon\s+social|emisor|proveedor)\s*[:\-]\s*([^\n\r]{3,80})", raw_text, flags=re.IGNORECASE)
    if m:
        return m.group(1).strip()
    return None


def _regex_extract(
    raw_text: str,
    file_name: str,
    fallback_hash: str,
    ocr_confidence: float | None = None,
    ocr_metrics: dict[str, object] | None = None,
) -> tuple[dict[str, object], dict[str, object]]:
    raw_text = _clean_text(raw_text)
    document_type, base_confidence = _infer_document_type(raw_text, file_name)

    rut_match = re.search(r"\bcl\s*tin\s*([0-9\.\-kK]{8,16})\b", raw_text, flags=re.IGNORECASE)
    if rut_match and rut_match.lastindex:
        rut_value = rut_match.group(1)
    else:
        rut_value = None

    rut_match = re.search(r"\b\d{1,2}(?:\.?\d{3}){2}-[\dkK]\b", raw_text) if not rut_value else None
    if not rut_match:
        rut_match = re.search(r"\b\d{7,8}-[\dkK]\b", raw_text) if not rut_value else None

    folio_match = re.search(r"(?:folio|n[ro\.]?|numero)\s*[:#-]?\s*([0-9]{4,})", raw_text, flags=re.IGNORECASE)
    parsed_issue_date = parse_issue_date(raw_text)

    total_match = None
    for pattern in [
        r"(?:monto\s+total|total\s+a\s+pagar|total)\s*[:$]?\s*([0-9\.,]+)",
        r"\btotal\b[^0-9]{0,8}([0-9\.,]+)",
    ]:
        total_match = re.search(pattern, raw_text, flags=re.IGNORECASE)
        if total_match:
            break

    total_value = None
    if total_match and total_match.lastindex:
        total_value = _parse_amount(total_match.group(1))
    else:
        amount_candidates: list[float] = []
        for candidate in re.findall(r"\b\d{1,3}(?:[\.,]\d{3})+(?:[\.,]\d{1,2})?\b", raw_text):
            parsed = _parse_amount(candidate)
            if parsed is not None:
                amount_candidates.append(parsed)
        if amount_candidates:
            total_value = max(amount_candidates)

    supplier_name = _extract_supplier_name(raw_text)

    evidence_count = sum(
        [
            1 if document_type != "unknown" else 0,
            1 if (rut_value or rut_match) else 0,
            1 if folio_match else 0,
            1 if parsed_issue_date else 0,
            1 if total_value is not None else 0,
        ]
    )
    confidence = min(0.99, base_confidence + (0.01 * evidence_count))
    if ocr_confidence is not None:
        ocr_score = float(ocr_confidence)
        if ocr_score >= 0.7:
            confidence = min(0.99, (confidence * 0.7) + (ocr_score * 0.3))
        else:
            confidence = max(confidence, min(0.99, (confidence * 0.9) + (ocr_score * 0.1)))

    text_for_country = f"{raw_text}\n{file_name}".lower()
    text_for_country_normalized = re.sub(r"[_\-]", " ", text_for_country)
    cl_markers = ["rut", "sii", "dte", "boleta", "clp", "servicio de impuestos internos"]
    intl_markers = ["google play", "youtube", "openai", "github", "invoice", "receipt"]
    has_cl_marker = any(re.search(rf"\b{re.escape(marker)}\b", text_for_country_normalized) for marker in cl_markers)
    if rut_value or rut_match or has_cl_marker:
        inferred_country = "CL"
    elif any(marker in text_for_country for marker in intl_markers):
        inferred_country = "INTL"
    else:
        inferred_country = "unknown"

    supplier_tax_id = _format_rut(rut_value) if rut_value else (_format_rut(rut_match.group(0)) if rut_match else None)
    if inferred_country != "CL" and not supplier_tax_id:
        supplier_tax_id = "FOREIGN"

    classification = {
        "document_type": document_type,
        "country_code": inferred_country,
        "confidence": confidence,
    }

    extraction = {
        "supplier_name": supplier_name or "Proveedor no identificado",
        "supplier_tax_id": supplier_tax_id,
        "document_number": folio_match.group(1) if folio_match and folio_match.lastindex else fallback_hash[:8],
        "issue_date": parsed_issue_date or _utc_today(),
        "currency": _normalize_currency(raw_text),
        "subtotal": None,
        "tax": None,
        "total": total_value,
        "amount_due": total_value,
        "description": "Gasto tributario extraido",
        "extra_fields": {
            "source_file": file_name,
            "rescued_raw_hash": fallback_hash,
            "raw_text_preview": raw_text[:1000],
            "ocr_confidence": float(ocr_confidence or 0.0),
            "ocr_metrics": ocr_metrics or {},
        },
    }
    return classification, extraction


def _infer_concept(extraction: dict[str, object]) -> str:
    text = f"{extraction.get('description', '')} {extraction.get('supplier_name', '')}".lower()
    keyword_rules = [
        ("bencina", "Combustible"),
        ("estacion", "Combustible"),
        ("servicio", "Servicios"),
        ("hotel", "Viaticos"),
        ("taxi", "Transporte"),
        ("uber", "Transporte"),
        ("comida", "Alimentacion"),
        ("restaurant", "Alimentacion"),
        ("office", "Insumos"),
        ("papeleria", "Insumos"),
    ]
    for key, concept in keyword_rules:
        if key in text:
            return concept
    return "Gasto Operacional"


def _extract_document_data(document_row: sqlite3.Row) -> tuple[dict[str, object], dict[str, object], bool]:
    raw_text = ""
    ocr_confidence = 0.0
    ocr_metrics: dict[str, object] = {}

    ocr_module = _load_back_ocr_module()
    if ocr_module is not None:
        try:
            ocr_result = ocr_module.extract_document_text(
                storage_path=document_row["storage_path"],
                mime_type=document_row["mime_type"] or "",
                file_name=document_row["file_name"] or "",
            )
            raw_text = str(ocr_result.get("text") or "")
            ocr_confidence = float(ocr_result.get("confidence") or 0.0)
            ocr_metrics = dict(ocr_result.get("metrics") or {})
        except Exception:
            raw_text = ""

    if not raw_text and (document_row["mime_type"] or "") == "application/pdf":
        raw_text = _extract_text_from_pdf(document_row["storage_path"])
        ocr_metrics = {"engine": "pypdf-fallback", "char_count": len(raw_text.strip())}
        ocr_confidence = 0.45 if raw_text else 0.0

    if not raw_text and (document_row["mime_type"] or "") == "message/rfc822":
        raw_text = _extract_text_from_eml(document_row["storage_path"])
        ocr_metrics = {"engine": "eml-parser", "char_count": len(raw_text.strip())}
        ocr_confidence = 0.75 if raw_text else 0.0

    if raw_text.strip():
        classification, extraction = _regex_extract(
            raw_text=raw_text,
            file_name=document_row["file_name"] or "",
            fallback_hash=document_row["document_hash"] or "",
            ocr_confidence=ocr_confidence,
            ocr_metrics=ocr_metrics,
        )
    else:
        classification, extraction = _regex_extract(
            raw_text=document_row["file_name"] or "",
            file_name=document_row["file_name"] or "",
            fallback_hash=document_row["document_hash"] or "",
            ocr_confidence=ocr_confidence,
            ocr_metrics=ocr_metrics,
        )

    intelligent_requires_review = False
    intelligence_module = _load_back_intelligence_module()
    ai_candidate: dict[str, object] | None = None

    ollama_module = _load_back_ollama_module()
    if ollama_module is not None and raw_text.strip():
        try:
            ai_candidate = ollama_module.fetch_ai_candidate(
                content=raw_text,
                file_name=document_row["file_name"] or "",
                mime_type=document_row["mime_type"] or "",
            )
        except Exception:
            ai_candidate = None

    if intelligence_module is not None:
        try:
            routed = intelligence_module.route_extraction(
                tenant_id=document_row["tenant_id"] or "default",
                regex_classification=classification,
                regex_extraction=extraction,
                ocr_confidence=ocr_confidence,
                ai_candidate=ai_candidate,
            )
            classification = dict(routed.get("classification") or classification)
            extraction = dict(routed.get("extraction") or extraction)
            intelligent_requires_review = bool(routed.get("requires_review", False))

            extra_fields = dict(extraction.get("extra_fields") or {})
            extra_fields["intelligent_routing"] = dict(routed.get("routing") or {})
            extraction["extra_fields"] = extra_fields
        except Exception:
            pass

    return classification, extraction, intelligent_requires_review


def _validate_extraction(classification: dict[str, object], extraction: dict[str, object]) -> tuple[list[dict[str, str]], bool]:
    warnings: list[dict[str, str]] = []
    requires_review = False

    confidence = float(classification.get("confidence") or 0.0)
    country = str(classification.get("country_code") or "unknown")
    doc_type = str(classification.get("document_type") or "unknown")

    if confidence < 0.75:
        requires_review = True
        _add_custom_warning(
            warnings,
            code="LOW_CONFIDENCE",
            field_name="classification",
            message="Documento requiere revision manual por baja confianza",
        )

    if doc_type == "unknown":
        requires_review = True
        _add_custom_warning(
            warnings,
            code="UNKNOWN_DOC_TYPE",
            field_name="document_type",
            message="No se pudo determinar el tipo documental",
        )

    supplier_tax_id = str(extraction.get("supplier_tax_id") or "")
    if country == "CL" and not supplier_tax_id:
        _add_custom_warning(
            warnings,
            code="MISSING_TAX_ID",
            field_name="RUT",
            message="No se detecto RUT del emisor",
        )

    subtotal = extraction.get("subtotal")
    tax = extraction.get("tax")
    total = extraction.get("total")
    if subtotal is not None and tax is not None and total is not None:
        try:
            expected = float(subtotal) + float(tax)
            if abs(expected - float(total)) > 1.0:
                requires_review = True
                _add_custom_warning(
                    warnings,
                    code="TOTAL_INCONSISTENT",
                    field_name="Total",
                    message="Inconsistencia contable: subtotal + impuesto no coincide con total",
                )
        except (TypeError, ValueError):
            pass

    return warnings, requires_review


def _build_mapping(document_row: sqlite3.Row, extraction: dict[str, object]) -> tuple[dict[str, str], dict[str, str], list[str], list[dict[str, str]]]:
    missing_fields: list[str] = []
    warnings: list[dict[str, str]] = []

    amount_due = extraction.get("amount_due") or extraction.get("total")
    period = document_row["period"] or (str(extraction.get("issue_date"))[:7] if extraction.get("issue_date") else "")
    supplier_rut = str(extraction.get("supplier_tax_id") or "")
    concept = _infer_concept(extraction)
    center_cost = document_row["center_cost"] or DEFAULT_CENTER_COST
    observations = "Captura automatica"

    header: dict[str, str] = {
        "Responsable": document_row["responsible"] or "",
        "Rut": COMPANY_RUT_DEFAULT,
        "Periodo": period,
        "Monto a pagar": str(amount_due) if amount_due is not None else "",
        "Fecha": _utc_today(),
        "Autorizado por": DEFAULT_APPROVER,
    }

    detail: dict[str, str] = {
        "Mes": month_number(str(extraction.get("issue_date") or ""), document_row["period"]),
        "Cuenta": DEFAULT_ACCOUNT,
        "Descripcion Cuenta": DEFAULT_ACCOUNT_DESCRIPTION,
        "Fecha": str(extraction.get("issue_date") or ""),
        "Nro Boleta Factura": str(extraction.get("document_number") or ""),
        "Descripcion del gasto": str(extraction.get("description") or ""),
        "Concepto": concept,
        "Centro Costo": center_cost,
        "PROVEEDOR": str(extraction.get("supplier_name") or ""),
        "RUT": supplier_rut,
        "Observaciones": observations,
        "Total": str(extraction.get("total") or extraction.get("amount_due") or ""),
    }

    for field_name, value in header.items():
        if not value:
            missing_fields.append(field_name)
            _add_warning(warnings, field_name)

    for field_name, value in detail.items():
        if not value:
            missing_fields.append(field_name)
            _add_warning(warnings, field_name)

    return header, detail, missing_fields, warnings


def _persist_audit_event(
    conn: sqlite3.Connection,
    *,
    entity_type: str,
    entity_id: str,
    event_type: str,
    payload: dict[str, object],
    actor_id: str | None = None,
) -> None:
    payload_with_meta = dict(payload)
    payload_with_meta["correlation_id"] = get_correlation_id()
    conn.execute(
        """
        INSERT INTO audit_events (
            id, entity_type, entity_id, event_type, payload_json, actor_id, created_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        (
            str(uuid.uuid4()),
            entity_type,
            entity_id,
            event_type,
            json.dumps(payload_with_meta, ensure_ascii=True),
            actor_id,
            _utc_now_iso(),
        ),
    )


def _apply_field_overrides(conn: sqlite3.Connection, document_id: str, fields: dict[str, str | int | float | None]) -> int:
    mapping_row = conn.execute(
        """
        SELECT header_fields_json, detail_fields_json
        FROM document_template_mapping
        WHERE document_id = ?
        """,
        (document_id,),
    ).fetchone()
    if not mapping_row:
        raise HTTPException(status_code=409, detail="Document must be processed before override")

    header = json.loads(mapping_row["header_fields_json"])
    detail = json.loads(mapping_row["detail_fields_json"])

    for field_name, value in fields.items():
        text_value = "" if value is None else str(value)
        if field_name in header:
            header[field_name] = text_value
        if field_name in detail:
            detail[field_name] = text_value

    missing_fields = [k for k, v in {**header, **detail}.items() if not v]
    warnings = []
    for field_name in missing_fields:
        _add_warning(warnings, field_name)

    conn.execute(
        """
        UPDATE document_template_mapping
        SET header_fields_json = ?, detail_fields_json = ?, missing_fields_json = ?, warnings_count = ?
        WHERE document_id = ?
        """,
        (
            json.dumps(header, ensure_ascii=True),
            json.dumps(detail, ensure_ascii=True),
            json.dumps(missing_fields, ensure_ascii=True),
            len(warnings),
            document_id,
        ),
    )
    conn.execute(
        "UPDATE documents SET warnings_json = ?, updated_at = ? WHERE id = ?",
        (json.dumps(warnings, ensure_ascii=True), _utc_now_iso(), document_id),
    )
    return len(fields)


def _write_rendition_file(*, rendition_id: str, header: dict[str, str], details: list[dict[str, str]]) -> str:
    try:
        from openpyxl import Workbook
    except ModuleNotFoundError as exc:
        raise HTTPException(status_code=500, detail="Missing dependency openpyxl") from exc

    export_dir = Path(EXPORT_DIR)
    export_dir.mkdir(parents=True, exist_ok=True)
    file_path = export_dir / f"rendition_{rendition_id}.xlsx"

    def normalize_label(value: str) -> str:
        ascii_text = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode("ascii")
        ascii_text = ascii_text.replace("\n", " ")
        ascii_text = re.sub(r"\s+", " ", ascii_text).strip().lower()
        return ascii_text

    template_path = Path(TEMPLATE_XLSX_PATH)
    if template_path.exists():
        from openpyxl import load_workbook

        wb = load_workbook(template_path)
        ws = wb.active

        header_targets = {normalize_label(k): k for k in HEADER_FIELDS}
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                found = normalize_label(str(cell.value))
                if found in header_targets:
                    field_name = header_targets[found]
                    ws.cell(row=cell.row, column=cell.column + 1, value=header.get(field_name, ""))

        best_row = None
        best_map: dict[str, int] = {}
        detail_targets = {normalize_label(k): k for k in DETAIL_FIELDS}
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            row_map: dict[str, int] = {}
            for cell in row:
                if cell.value is None:
                    continue
                found = normalize_label(str(cell.value))
                if found in detail_targets:
                    row_map[detail_targets[found]] = cell.column
            if len(row_map) > len(best_map):
                best_map = row_map
                best_row = row[0].row

        if best_row and len(best_map) >= 6:
            start_row = best_row + 1
            for offset, item in enumerate(details):
                row_number = start_row + offset
                for field_name, col_number in best_map.items():
                    ws.cell(row=row_number, column=col_number, value=item.get(field_name, ""))
        else:
            row_cursor = ws.max_row + 2
            for idx, field_name in enumerate(DETAIL_FIELDS, start=1):
                ws.cell(row=row_cursor, column=idx, value=field_name)
            for item in details:
                row_cursor += 1
                for idx, field_name in enumerate(DETAIL_FIELDS, start=1):
                    ws.cell(row=row_cursor, column=idx, value=item.get(field_name, ""))
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Rendicion"

        ws["A1"] = "RENDICION GASTOS"
        row_cursor = 3
        for field_name in HEADER_FIELDS:
            ws.cell(row=row_cursor, column=1, value=field_name)
            ws.cell(row=row_cursor, column=2, value=header.get(field_name, ""))
            row_cursor += 1

        row_cursor += 1
        for idx, field_name in enumerate(DETAIL_FIELDS, start=1):
            ws.cell(row=row_cursor, column=idx, value=field_name)

        for item in details:
            row_cursor += 1
            for idx, field_name in enumerate(DETAIL_FIELDS, start=1):
                ws.cell(row=row_cursor, column=idx, value=item.get(field_name, ""))

    wb.save(file_path)
    return str(file_path)


async def upload_document(
    file: UploadFile,
    tenant_id: str,
    responsible: str | None,
    period: str | None,
    center_cost: str | None,
) -> dict[str, object]:
    if file.content_type not in ALLOWED_MIME_TYPES:
        raise HTTPException(status_code=415, detail="Unsupported mime type")

    safe_file_name = _sanitize_filename(file.filename)
    if "\x00" in safe_file_name:
        raise HTTPException(status_code=400, detail="Invalid file name")

    extension = Path(safe_file_name).suffix.lower()
    if extension and extension not in ALLOWED_EXTENSIONS:
        raise HTTPException(status_code=415, detail="Unsupported file extension")

    expected_mime = EXPECTED_MIME_BY_EXTENSION.get(extension)
    if expected_mime and file.content_type != expected_mime:
        _audit_upload_blocked(tenant_id=tenant_id, file_name=safe_file_name, reason="mime_extension_mismatch")
        raise HTTPException(status_code=415, detail="Mime type does not match file extension")

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Empty file")
    if len(content) > MAX_UPLOAD_BYTES:
        raise HTTPException(status_code=413, detail="File too large")

    signature_extension = _detect_file_signature(content)
    canonical_extension = _canonical_extension(extension)
    canonical_signature = _canonical_extension(signature_extension)
    if signature_extension is None or (extension in EXPECTED_MIME_BY_EXTENSION and canonical_signature != canonical_extension):
        reason = "invalid_or_mismatched_file_signature"
        _audit_upload_blocked(tenant_id=tenant_id, file_name=safe_file_name, reason=reason)
        if UPLOAD_QUARANTINE_ENABLED:
            _quarantine_upload(safe_file_name=safe_file_name, content=content, reason=reason)
            raise HTTPException(status_code=422, detail="File moved to quarantine")
        raise HTTPException(status_code=422, detail="Invalid file signature")

    document_hash = hashlib.sha256(content).hexdigest()
    now = _utc_now_iso()

    conn = get_db_connection()
    try:
        existing = conn.execute(
            """
            SELECT id, warnings_json
            FROM documents
            WHERE tenant_id = ? AND document_hash = ?
            """,
            (tenant_id, document_hash),
        ).fetchone()

        if existing:
            return {
                "document_id": existing["id"],
                "status": "RECEIVED",
                "duplicate": True,
                "warnings": json.loads(existing["warnings_json"]),
            }

        document_id = str(uuid.uuid4())
        storage_dir = Path(STORAGE_DIR)
        storage_dir.mkdir(parents=True, exist_ok=True)
        extension = Path(safe_file_name).suffix or ".bin"
        storage_path = storage_dir / f"{document_id}{extension}"
        storage_path.write_bytes(content)

        conn.execute(
            """
            INSERT INTO documents (
                id, tenant_id, document_hash, file_name, mime_type,
                status, responsible, period, center_cost, storage_path,
                warnings_json, created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                document_id,
                tenant_id,
                document_hash,
                safe_file_name,
                file.content_type or "application/octet-stream",
                "RECEIVED",
                responsible,
                period,
                center_cost,
                str(storage_path),
                "[]",
                now,
                now,
            ),
        )
        _persist_audit_event(
            conn,
            entity_type="document",
            entity_id=document_id,
            event_type="DOCUMENT_RECEIVED",
            payload={"tenant_id": tenant_id, "file_name": safe_file_name},
        )
        conn.commit()

        # Publica el documento para procesamiento asincrono.
        enqueued = enqueue_document_processing(document_id)
        return {
            "document_id": document_id,
            "status": "RECEIVED",
            "duplicate": False,
            "enqueued": enqueued,
            "warnings": [],
        }
    finally:
        conn.close()


def get_document(document_id: str) -> dict[str, object]:
    conn = get_db_connection()
    try:
        row = conn.execute(
            """
            SELECT id, status, warnings_json
            FROM documents
            WHERE id = ?
            """,
            (document_id,),
        ).fetchone()

        extraction_row = conn.execute(
            """
            SELECT document_type, country_code, confidence, extraction_json
            FROM document_extractions
            WHERE document_id = ?
            """,
            (document_id,),
        ).fetchone()
    finally:
        conn.close()

    if not row:
        raise HTTPException(status_code=404, detail="Document not found")

    classification = {
        "document_type": "unknown",
        "country_code": "unknown",
        "confidence": 0.0,
    }
    extraction_payload: dict[str, object] = {
        "supplier_name": None,
        "supplier_tax_id": None,
        "document_number": None,
        "issue_date": None,
        "currency": None,
        "total": None,
    }

    if extraction_row:
        classification = {
            "document_type": extraction_row["document_type"] or "unknown",
            "country_code": extraction_row["country_code"] or "unknown",
            "confidence": float(extraction_row["confidence"] or 0.0),
        }
        extraction_payload = json.loads(extraction_row["extraction_json"])

    return {
        "document_id": row["id"],
        "status": row["status"],
        "classification": classification,
        "extraction": extraction_payload,
        "warnings": json.loads(row["warnings_json"]),
        "review_required": row["status"] == "REVIEW_REQUIRED",
    }


def process_document(document_id: str) -> dict[str, object]:
    conn = get_db_connection()
    try:
        row = conn.execute(
            """
            SELECT id, file_name, tenant_id, document_hash, mime_type, storage_path,
                   responsible, period, center_cost
            FROM documents
            WHERE id = ?
            """,
            (document_id,),
        ).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Document not found")

        conn.execute(
            "UPDATE documents SET status = ?, updated_at = ? WHERE id = ?",
            ("PROCESSING", _utc_now_iso(), document_id),
        )

        classification, extraction, intelligent_requires_review = _extract_document_data(row)
        header, detail, missing_fields, warnings = _build_mapping(row, extraction)

        validation_warnings, review_required = _validate_extraction(classification, extraction)
        review_required = review_required or intelligent_requires_review
        warnings.extend(validation_warnings)
        final_status = "REVIEW_REQUIRED" if review_required else "COMPLETED"

        conn.execute(
            """
            INSERT INTO document_extractions (
                id, document_id, provider, document_type, country_code, confidence,
                extraction_json, raw_extraction_json, extra_fields_json, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(document_id) DO UPDATE SET
                provider=excluded.provider,
                document_type=excluded.document_type,
                country_code=excluded.country_code,
                confidence=excluded.confidence,
                extraction_json=excluded.extraction_json,
                raw_extraction_json=excluded.raw_extraction_json,
                extra_fields_json=excluded.extra_fields_json
            """,
            (
                str(uuid.uuid4()),
                document_id,
                "regex-pdf-parser",
                classification["document_type"],
                classification["country_code"],
                classification["confidence"],
                json.dumps(extraction, ensure_ascii=True),
                json.dumps(extraction, ensure_ascii=True),
                json.dumps(extraction.get("extra_fields", {}), ensure_ascii=True),
                _utc_now_iso(),
            ),
        )

        conn.execute(
            """
            INSERT INTO document_template_mapping (
                id, document_id, header_fields_json, detail_fields_json,
                missing_fields_json, warnings_count, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(document_id) DO UPDATE SET
                header_fields_json=excluded.header_fields_json,
                detail_fields_json=excluded.detail_fields_json,
                missing_fields_json=excluded.missing_fields_json,
                warnings_count=excluded.warnings_count
            """,
            (
                str(uuid.uuid4()),
                document_id,
                json.dumps(header, ensure_ascii=True),
                json.dumps(detail, ensure_ascii=True),
                json.dumps(missing_fields, ensure_ascii=True),
                len(warnings),
                _utc_now_iso(),
            ),
        )

        conn.execute(
            "UPDATE documents SET status = ?, warnings_json = ?, updated_at = ? WHERE id = ?",
            (final_status, json.dumps(warnings, ensure_ascii=True), _utc_now_iso(), document_id),
        )

        if final_status == "REVIEW_REQUIRED":
            now = _utc_now_iso()
            conn.execute(
                """
                INSERT INTO review_tasks (
                    id, document_id, status, reason, decision, reviewer_id,
                    resolution_reason, resolved_at, warnings_json, created_at, updated_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(document_id) DO UPDATE SET
                    status=excluded.status,
                    reason=excluded.reason,
                    decision=excluded.decision,
                    reviewer_id=excluded.reviewer_id,
                    resolution_reason=excluded.resolution_reason,
                    resolved_at=excluded.resolved_at,
                    warnings_json=excluded.warnings_json,
                    updated_at=excluded.updated_at
                """,
                (
                    str(uuid.uuid4()),
                    document_id,
                    "PENDING",
                    "Validaciones requieren revision manual",
                    None,
                    None,
                    None,
                    None,
                    json.dumps(warnings, ensure_ascii=True),
                    now,
                    now,
                ),
            )
        else:
            conn.execute("DELETE FROM review_tasks WHERE document_id = ?", (document_id,))

        _persist_audit_event(
            conn,
            entity_type="document",
            entity_id=document_id,
            event_type="EXTRACTION_COMPLETED",
            payload={
                "status": final_status,
                "document_type": classification["document_type"],
                "missing_fields": missing_fields,
            },
        )
        conn.commit()
    finally:
        conn.close()

    return {
        "document_id": document_id,
        "status": final_status,
        "classification": classification,
        "warnings_count": len(warnings),
        "missing_fields": missing_fields,
    }


def override_document_fields(document_id: str, request: OverrideRequest) -> dict[str, object]:
    conn = get_db_connection()
    try:
        row = conn.execute(
            "SELECT id FROM documents WHERE id = ?",
            (document_id,),
        ).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Document not found")

        applied = _apply_field_overrides(conn, document_id, request.fields)

        _persist_audit_event(
            conn,
            entity_type="document",
            entity_id=document_id,
            event_type="OVERRIDE_APPLIED",
            payload={"fields": request.fields, "reason": request.reason},
        )
        conn.commit()
    finally:
        conn.close()

    return {
        "document_id": document_id,
        "status": "COMPLETED",
        "overrides_applied": applied,
    }


def _review_row_or_404(conn, document_id: str):
    review_row = conn.execute(
        """
        SELECT document_id, status, reason, decision, reviewer_id, resolution_reason, resolved_at,
               warnings_json, created_at, updated_at
        FROM review_tasks
        WHERE document_id = ?
        """,
        (document_id,),
    ).fetchone()
    if not review_row:
        raise HTTPException(status_code=404, detail="Review task not found")
    return review_row


def list_pending_reviews() -> dict[str, object]:
    conn = get_db_connection()
    try:
        rows = conn.execute(
            """
            SELECT r.document_id, r.reason, r.warnings_json, r.reviewer_id,
                   d.file_name, d.tenant_id, d.updated_at
            FROM review_tasks r
            JOIN documents d ON d.id = r.document_id
            WHERE r.status = 'PENDING'
            ORDER BY d.updated_at DESC
            """
        ).fetchall()
    finally:
        conn.close()

    return {
        "total": len(rows),
        "items": [
            {
                "document_id": row["document_id"],
                "tenant_id": row["tenant_id"],
                "file_name": row["file_name"],
                "reason": row["reason"],
                "reviewer_id": row["reviewer_id"],
                "warnings": json.loads(row["warnings_json"]),
                "updated_at": row["updated_at"],
            }
            for row in rows
        ],
    }


def get_review_detail(document_id: str) -> dict[str, object]:
    conn = get_db_connection()
    try:
        review_row = _review_row_or_404(conn, document_id)
        document_row = conn.execute(
            """
            SELECT id, tenant_id, file_name, status, warnings_json, updated_at
            FROM documents
            WHERE id = ?
            """,
            (document_id,),
        ).fetchone()
        extraction_row = conn.execute(
            """
            SELECT document_type, country_code, confidence, extraction_json
            FROM document_extractions
            WHERE document_id = ?
            """,
            (document_id,),
        ).fetchone()
        mapping_row = conn.execute(
            """
            SELECT header_fields_json, detail_fields_json, missing_fields_json
            FROM document_template_mapping
            WHERE document_id = ?
            """,
            (document_id,),
        ).fetchone()
        audit_rows = conn.execute(
            """
            SELECT event_type, payload_json, actor_id, created_at
            FROM audit_events
            WHERE entity_type = 'document' AND entity_id = ?
            ORDER BY created_at ASC
            """,
            (document_id,),
        ).fetchall()
    finally:
        conn.close()

    if not document_row:
        raise HTTPException(status_code=404, detail="Document not found")

    manual_changes = []
    for row in audit_rows:
        if row["event_type"] in {"OVERRIDE_APPLIED", "REVIEW_OVERRIDE_APPLIED", "REVIEW_RESOLVED"}:
            manual_changes.append(
                {
                    "event_type": row["event_type"],
                    "actor_id": row["actor_id"],
                    "created_at": row["created_at"],
                    "payload": json.loads(row["payload_json"]),
                }
            )

    return {
        "document": {
            "document_id": document_row["id"],
            "tenant_id": document_row["tenant_id"],
            "file_name": document_row["file_name"],
            "status": document_row["status"],
            "warnings": json.loads(document_row["warnings_json"]),
            "updated_at": document_row["updated_at"],
        },
        "review": {
            "status": review_row["status"],
            "reason": review_row["reason"],
            "decision": review_row["decision"],
            "reviewer_id": review_row["reviewer_id"],
            "resolution_reason": review_row["resolution_reason"],
            "resolved_at": review_row["resolved_at"],
            "warnings": json.loads(review_row["warnings_json"]),
            "created_at": review_row["created_at"],
            "updated_at": review_row["updated_at"],
        },
        "classification": {
            "document_type": extraction_row["document_type"] if extraction_row else "unknown",
            "country_code": extraction_row["country_code"] if extraction_row else "unknown",
            "confidence": float(extraction_row["confidence"] or 0.0) if extraction_row else 0.0,
        },
        "extraction": json.loads(extraction_row["extraction_json"]) if extraction_row else {},
        "mapping": {
            "header": json.loads(mapping_row["header_fields_json"]) if mapping_row else {},
            "detail": json.loads(mapping_row["detail_fields_json"]) if mapping_row else {},
            "missing_fields": json.loads(mapping_row["missing_fields_json"]) if mapping_row else [],
        },
        "manual_changes": manual_changes,
    }


def apply_review_overrides(document_id: str, request: ReviewOverridesRequest) -> dict[str, object]:
    conn = get_db_connection()
    try:
        review_row = _review_row_or_404(conn, document_id)
        if str(review_row["status"]).startswith("RESOLVED"):
            raise HTTPException(status_code=409, detail="Review task already resolved")

        applied = _apply_field_overrides(conn, document_id, request.overrides)
        now = _utc_now_iso()

        conn.execute(
            """
            UPDATE review_tasks
            SET reviewer_id = ?, reason = ?, updated_at = ?
            WHERE document_id = ?
            """,
            (request.reviewer_id, request.reason, now, document_id),
        )
        _persist_audit_event(
            conn,
            entity_type="document",
            entity_id=document_id,
            event_type="REVIEW_OVERRIDE_APPLIED",
            payload={"reason": request.reason, "fields": request.overrides, "applied": applied},
            actor_id=request.reviewer_id,
        )
        conn.commit()
    finally:
        conn.close()

    return {
        "document_id": document_id,
        "status": "PENDING",
        "overrides_applied": applied,
        "reviewer_id": request.reviewer_id,
    }


def _finalize_review(document_id: str, decision: str, request: ReviewActionRequest) -> dict[str, object]:
    conn = get_db_connection()
    try:
        review_row = _review_row_or_404(conn, document_id)
        if str(review_row["status"]).startswith("RESOLVED"):
            raise HTTPException(status_code=409, detail="Review task already resolved")

        applied = 0
        if request.overrides:
            applied = _apply_field_overrides(conn, document_id, request.overrides)

        new_doc_status = "COMPLETED" if decision == "approve" else "REJECTED"
        new_review_status = "RESOLVED_APPROVED" if decision == "approve" else "RESOLVED_REJECTED"
        now = _utc_now_iso()

        conn.execute(
            "UPDATE documents SET status = ?, updated_at = ? WHERE id = ?",
            (new_doc_status, now, document_id),
        )
        conn.execute(
            """
            UPDATE review_tasks
            SET status = ?, decision = ?, reviewer_id = ?, resolution_reason = ?, resolved_at = ?, updated_at = ?
            WHERE document_id = ?
            """,
            (new_review_status, decision, request.reviewer_id, request.reason, now, now, document_id),
        )

        _persist_audit_event(
            conn,
            entity_type="document",
            entity_id=document_id,
            event_type="REVIEW_RESOLVED",
            payload={
                "decision": decision,
                "reason": request.reason,
                "overrides_applied": applied,
                "overrides": request.overrides,
            },
            actor_id=request.reviewer_id,
        )
        conn.commit()
    finally:
        conn.close()

    return {
        "document_id": document_id,
        "decision": decision,
        "status": new_doc_status,
        "reviewer_id": request.reviewer_id,
        "resolved_at": now,
        "overrides_applied": applied,
    }


def approve_review(document_id: str, request: ReviewActionRequest) -> dict[str, object]:
    return _finalize_review(document_id, "approve", request)


def reject_review(document_id: str, request: ReviewActionRequest) -> dict[str, object]:
    return _finalize_review(document_id, "reject", request)


def resolve_review(document_id: str, request: ReviewResolveRequest) -> dict[str, object]:
    decision = request.decision.strip().lower()
    if decision not in {"approve", "reject"}:
        raise HTTPException(status_code=400, detail="decision must be approve or reject")

    action_request = ReviewActionRequest(
        reviewer_id=request.reviewer_id,
        reason=request.reason,
        overrides=request.overrides,
    )
    if decision == "approve":
        return approve_review(document_id, action_request)
    return reject_review(document_id, action_request)


def generate_rendition(request: RenditionGenerateRequest) -> dict[str, object]:
    if not request.document_ids:
        raise HTTPException(status_code=400, detail="document_ids must not be empty")

    conn = get_db_connection()
    try:
        mappings: list[sqlite3.Row] = []
        for document_id in request.document_ids:
            mapping_row = conn.execute(
                """
                SELECT m.detail_fields_json, m.header_fields_json, m.missing_fields_json,
                       d.tenant_id
                FROM document_template_mapping m
                JOIN documents d ON d.id = m.document_id
                WHERE m.document_id = ?
                """,
                (document_id,),
            ).fetchone()
            if not mapping_row:
                raise HTTPException(status_code=409, detail=f"Document {document_id} has no mapping")
            if mapping_row["tenant_id"] != request.tenant_id:
                raise HTTPException(status_code=400, detail=f"Document {document_id} belongs to another tenant")
            mappings.append(mapping_row)

        rendition_id = str(uuid.uuid4())
        details: list[dict[str, str]] = []
        missing_all: list[str] = []
        warnings_count = 0

        first_header = json.loads(mappings[0]["header_fields_json"])
        first_header["Periodo"] = request.period

        for row in mappings:
            detail = json.loads(row["detail_fields_json"])
            details.append(detail)
            missing_fields = json.loads(row["missing_fields_json"])
            missing_all.extend(missing_fields)
            warnings_count += len(missing_fields)

        file_path = _write_rendition_file(rendition_id=rendition_id, header=first_header, details=details)
        summary = {
            "rows": len(details),
            "warnings": warnings_count,
            "missing_fields": sorted(set(missing_all)),
        }

        now = _utc_now_iso()
        conn.execute(
            """
            INSERT INTO renditions (
                id, tenant_id, period, template_version, file_path,
                status, warnings_count, summary_json, created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                rendition_id,
                request.tenant_id,
                request.period,
                request.template_version,
                file_path,
                "GENERATED",
                warnings_count,
                json.dumps(summary, ensure_ascii=True),
                now,
                now,
            ),
        )

        for row_number, (document_id, detail) in enumerate(zip(request.document_ids, details), start=1):
            missing_fields_row = [k for k, v in detail.items() if not v]
            warnings_row = []
            for field_name in missing_fields_row:
                _add_warning(warnings_row, field_name)

            conn.execute(
                """
                INSERT INTO rendition_items (
                    id, rendition_id, document_id, row_number, fields_json, warnings_json
                ) VALUES (?, ?, ?, ?, ?, ?)
                """,
                (
                    str(uuid.uuid4()),
                    rendition_id,
                    document_id,
                    row_number,
                    json.dumps(detail, ensure_ascii=True),
                    json.dumps(warnings_row, ensure_ascii=True),
                ),
            )

            _persist_audit_event(
                conn,
                entity_type="document",
                entity_id=document_id,
                event_type="RENDITION_ITEM_INCLUDED",
                payload={
                    "rendition_id": rendition_id,
                    "row_number": row_number,
                    "template_version": request.template_version,
                },
            )

        _persist_audit_event(
            conn,
            entity_type="rendition",
            entity_id=rendition_id,
            event_type="RENDITION_GENERATED",
            payload=summary,
        )

        conn.commit()
    finally:
        conn.close()

    return {
        "rendition_id": rendition_id,
        "file_url": file_path,
        "warnings_count": warnings_count,
        "status": "GENERATED",
    }


def get_rendition(rendition_id: str) -> dict[str, object]:
    conn = get_db_connection()
    try:
        row = conn.execute(
            """
            SELECT id, status, file_path, summary_json
            FROM renditions
            WHERE id = ?
            """,
            (rendition_id,),
        ).fetchone()
    finally:
        conn.close()

    if not row:
        raise HTTPException(status_code=404, detail="Rendition not found")

    return {
        "rendition_id": row["id"],
        "status": row["status"],
        "file_url": row["file_path"],
        "summary": json.loads(row["summary_json"]),
    }


def generate_rendition_by_filter(request: RenditionGenerateByFilterRequest) -> dict[str, object]:
    conn = get_db_connection()
    try:
        query = [
            """
            SELECT d.id
            FROM documents d
            JOIN document_template_mapping m ON m.document_id = d.id
            WHERE d.tenant_id = ?
              AND d.status IN ('COMPLETED', 'REVIEW_REQUIRED')
            """
        ]
        params: list[object] = [request.tenant_id]

        if request.period:
            query.append("AND COALESCE(d.period, '') = ?")
            params.append(request.period)
        if request.responsible:
            query.append("AND COALESCE(d.responsible, '') = ?")
            params.append(request.responsible)
        if request.center_cost:
            query.append("AND COALESCE(d.center_cost, '') = ?")
            params.append(request.center_cost)

        query.append("ORDER BY d.created_at ASC")

        rows = conn.execute("\n".join(query), tuple(params)).fetchall()
    finally:
        conn.close()

    document_ids = [str(row["id"]) for row in rows]
    if not document_ids:
        raise HTTPException(status_code=404, detail="No documents found for provided filters")

    return generate_rendition(
        RenditionGenerateRequest(
            tenant_id=request.tenant_id,
            period=request.period,
            document_ids=document_ids,
            template_version=request.template_version,
        )
    )


def list_rendition_items(rendition_id: str) -> dict[str, object]:
    conn = get_db_connection()
    try:
        rendition_row = conn.execute(
            "SELECT id, template_version, period FROM renditions WHERE id = ?",
            (rendition_id,),
        ).fetchone()
        if not rendition_row:
            raise HTTPException(status_code=404, detail="Rendition not found")

        rows = conn.execute(
            """
            SELECT document_id, row_number, fields_json, warnings_json
            FROM rendition_items
            WHERE rendition_id = ?
            ORDER BY row_number ASC
            """,
            (rendition_id,),
        ).fetchall()
    finally:
        conn.close()

    return {
        "rendition_id": rendition_id,
        "template_version": rendition_row["template_version"],
        "period": rendition_row["period"],
        "total": len(rows),
        "items": [
            {
                "document_id": row["document_id"],
                "row_number": row["row_number"],
                "fields": json.loads(row["fields_json"]),
                "warnings": json.loads(row["warnings_json"]),
            }
            for row in rows
        ],
    }


def download_rendition(rendition_id: str) -> str:
    conn = get_db_connection()
    try:
        row = conn.execute(
            "SELECT file_path FROM renditions WHERE id = ?",
            (rendition_id,),
        ).fetchone()
    finally:
        conn.close()

    if not row:
        raise HTTPException(status_code=404, detail="Rendition not found")

    file_path = str(row["file_path"] or "")
    if not file_path or not Path(file_path).exists():
        raise HTTPException(status_code=404, detail="Rendition file not found")
    return file_path
